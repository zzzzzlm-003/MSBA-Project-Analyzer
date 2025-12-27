"""
项目分析系统 - 本地AI分析版本（无需OpenAI API）
提取PDF文本后，由AI助手直接分析
"""

import os
import json
import re
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import requests

# 配置
PROJECTS_DIR = Path("data/projects")
PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = Path("data/output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TEXTS_DIR = Path("data/project_texts")
TEXTS_DIR.mkdir(parents=True, exist_ok=True)


class GoogleDriveDownloader:
    """从Google Drive共享链接下载PDF文件"""
    
    @staticmethod
    def extract_file_id(share_link: str) -> Optional[str]:
        """从Google Drive共享链接中提取文件ID"""
        patterns = [
            r'/file/d/([a-zA-Z0-9_-]+)',
            r'id=([a-zA-Z0-9_-]+)',
            r'/([a-zA-Z0-9_-]{25,})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, share_link)
            if match:
                return match.group(1)
        return None
    
    @staticmethod
    def extract_folder_id(folder_link: str) -> Optional[str]:
        """从Google Drive文件夹链接中提取文件夹ID"""
        patterns = [
            r'/folders/([a-zA-Z0-9_-]+)',
            r'id=([a-zA-Z0-9_-]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, folder_link)
            if match:
                return match.group(1)
        return None
    
    @staticmethod
    def download_file(file_id: str, output_path: Path) -> bool:
        """下载Google Drive文件"""
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        
        try:
            session = requests.Session()
            response = session.get(f"https://drive.google.com/uc?id={file_id}", stream=True)
            
            if 'download_warning' in response.url or 'confirm' in response.url:
                confirm_match = re.search(r'confirm=([^&]+)', response.url)
                if confirm_match:
                    confirm_token = confirm_match.group(1)
                    download_url = f"https://drive.google.com/uc?export=download&id={file_id}&confirm={confirm_token}"
                    response = session.get(download_url, stream=True)
            
            if response.headers.get('Content-Type', '').startswith('text/html'):
                download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
                response = requests.get(download_url, stream=True, allow_redirects=True)
            
            if response.status_code == 200:
                output_path.parent.mkdir(parents=True, exist_ok=True)
                with open(output_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                print(f"✓ 下载成功: {output_path.name}")
                return True
            else:
                print(f"✗ 下载失败: {response.status_code} - {output_path.name}")
                return False
                
        except Exception as e:
            print(f"✗ 下载错误: {output_path.name} - {str(e)}")
            return False


class PDFExtractor:
    """提取PDF文本内容"""
    
    @staticmethod
    def extract_text(pdf_path: Path) -> str:
        """从PDF文件中提取文本"""
        try:
            # 尝试使用PyPDF2
            try:
                import PyPDF2
                with open(pdf_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
                return text
            except ImportError:
                # 如果PyPDF2不可用，尝试pdfplumber
                try:
                    import pdfplumber
                    text = ""
                    with pdfplumber.open(pdf_path) as pdf:
                        for page in pdf.pages:
                            page_text = page.extract_text()
                            if page_text:
                                text += page_text + "\n"
                    return text
                except ImportError:
                    print("错误: 需要安装 PyPDF2 或 pdfplumber")
                    return ""
        except Exception as e:
            print(f"✗ PDF解析错误: {pdf_path.name} - {str(e)}")
            return ""
    
    @staticmethod
    def extract_all_pdfs_to_texts(pdf_dir: Path = None) -> List[Dict]:
        """提取所有PDF文本并保存为文本文件"""
        if pdf_dir is None:
            pdf_dir = PROJECTS_DIR
        
        pdf_files = list(pdf_dir.glob("*.pdf"))
        if not pdf_files:
            print("未找到PDF文件")
            return []
        
        extracted_files = []
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"\n[{i}/{len(pdf_files)}] 提取文本: {pdf_file.name}")
            
            text = PDFExtractor.extract_text(pdf_file)
            if not text:
                print(f"✗ 无法提取文本: {pdf_file.name}")
                continue
            
            # 保存文本文件
            text_file = TEXTS_DIR / f"{pdf_file.stem}.txt"
            with open(text_file, 'w', encoding='utf-8') as f:
                f.write(text)
            
            extracted_files.append({
                "pdf_file": pdf_file.name,
                "text_file": text_file.name,
                "text_length": len(text),
                "text_preview": text[:500]  # 前500字符预览
            })
            
            print(f"✓ 文本已保存: {text_file.name} ({len(text)} 字符)")
        
        # 保存索引文件
        index_file = TEXTS_DIR / "index.json"
        with open(index_file, 'w', encoding='utf-8') as f:
            json.dump(extracted_files, f, ensure_ascii=False, indent=2)
        
        print(f"\n✓ 共提取 {len(extracted_files)} 个PDF的文本")
        print(f"✓ 文本文件保存在: {TEXTS_DIR}")
        print(f"✓ 索引文件: {index_file}")
        
        return extracted_files


class ExcelExporter:
    """导出数据到Excel"""
    
    @staticmethod
    def export_to_excel(projects: List[Dict], output_path: Path):
        """导出项目数据到Excel文件"""
        df = pd.DataFrame(projects)
        
        # 重新排列列的顺序，重要信息在前
        column_order = [
            "项目编号", "项目名称", "公司名称", "所处行业", 
            "应用场景", "公司用心程度", "预期成果", 
            "技能要求", "项目描述摘要", "源文件"
        ]
        
        # 只保留存在的列
        existing_columns = [col for col in column_order if col in df.columns]
        other_columns = [col for col in df.columns if col not in column_order]
        df = df[existing_columns + other_columns]
        
        # 使用openpyxl创建格式化的Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='项目分析')
            
            # 获取工作表
            worksheet = writer.sheets['项目分析']
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 设置列宽，但不超过50
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 设置行高
            worksheet.row_dimensions[1].height = 25
            for row in range(2, len(df) + 2):
                worksheet.row_dimensions[row].height = 20
            
            # 设置文本换行
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        print(f"✓ Excel文件已导出: {output_path}")
        return df


def download_from_folder_link(folder_link: str) -> List[Path]:
    """
    从Google Drive文件夹链接下载所有PDF
    注意：这需要手动提供文件链接，或者使用Google Drive API
    """
    print("\n" + "="*60)
    print("Google Drive文件夹下载说明")
    print("="*60)
    print("\n由于Google Drive文件夹需要认证，建议使用以下方法：")
    print("\n方法1：手动下载（推荐）")
    print("1. 在浏览器中打开文件夹链接")
    print("2. 选择所有PDF文件，下载到本地")
    print("3. 将PDF文件放到 data/projects/ 目录")
    print("4. 运行: python project_analyzer_local.py --extract")
    
    print("\n方法2：提供文件链接列表")
    print("1. 在文件夹中，右键每个PDF文件 -> 获取链接")
    print("2. 将所有链接添加到 project_analyzer_config.json")
    print("3. 运行程序自动下载")
    
    print("\n方法3：使用Google Drive API（需要配置）")
    print("需要创建Google Cloud项目并启用Drive API")
    
    return []


def main():
    """主函数"""
    import sys
    
    print("=" * 60)
    print("项目分析系统 - 本地AI版本")
    print("=" * 60)
    print("\n此版本无需OpenAI API，PDF文本提取后由AI助手直接分析")
    
    if len(sys.argv) > 1 and sys.argv[1] == "--extract":
        # 只提取文本
        print("\n提取PDF文本...")
        PDFExtractor.extract_all_pdfs_to_texts()
        print("\n✓ 文本提取完成！")
        print("\n下一步：")
        print("1. 查看 data/project_texts/ 目录中的文本文件")
        print("2. 将文本内容提供给AI助手进行分析")
        print("3. 或运行分析脚本逐个处理")
        return
    
    print("\n请选择操作模式:")
    print("1. 从Google Drive链接下载PDF（需要提供文件链接）")
    print("2. 提取本地PDF的文本内容（为AI分析做准备）")
    print("3. 查看已提取的文本文件列表")
    
    choice = input("\n请输入选择 (1/2/3): ").strip()
    
    if choice == "1":
        folder_link = input("\n请输入Google Drive文件夹链接: ").strip()
        if folder_link:
            download_from_folder_link(folder_link)
        
        # 也可以尝试从配置文件读取单个文件链接
        config_path = Path("project_analyzer_config.json")
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                links = config.get("google_drive_links", [])
                valid_links = [l for l in links if l and not l.startswith("在此处") and not l.startswith("例如")]
                
                if valid_links:
                    print(f"\n从配置文件找到 {len(valid_links)} 个链接")
                    downloader = GoogleDriveDownloader()
                    downloaded = []
                    
                    for i, link in enumerate(valid_links, 1):
                        print(f"\n[{i}/{len(valid_links)}] 处理: {link[:50]}...")
                        file_id = downloader.extract_file_id(link)
                        if file_id:
                            filename = f"project_{i:03d}.pdf"
                            output_path = PROJECTS_DIR / filename
                            if downloader.download_file(file_id, output_path):
                                downloaded.append(output_path)
                    
                    if downloaded:
                        print(f"\n✓ 成功下载 {len(downloaded)} 个文件")
                        extract_choice = input("\n是否立即提取文本？(y/n): ").strip().lower()
                        if extract_choice == 'y':
                            PDFExtractor.extract_all_pdfs_to_texts()
    
    elif choice == "2":
        PDFExtractor.extract_all_pdfs_to_texts()
    
    elif choice == "3":
        index_file = TEXTS_DIR / "index.json"
        if index_file.exists():
            with open(index_file, 'r', encoding='utf-8') as f:
                files = json.load(f)
            print(f"\n已提取 {len(files)} 个PDF的文本:")
            for i, file_info in enumerate(files, 1):
                print(f"\n{i}. {file_info['pdf_file']}")
                print(f"   文本文件: {file_info['text_file']}")
                print(f"   文本长度: {file_info['text_length']} 字符")
                print(f"   预览: {file_info['text_preview'][:100]}...")
        else:
            print("\n尚未提取任何文本文件，请先运行选项2")


if __name__ == "__main__":
    main()


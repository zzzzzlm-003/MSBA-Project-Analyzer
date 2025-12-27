"""
直接分析PDF项目 - 无需OpenAI，由AI助手直接分析
"""

import json
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 配置
PROJECTS_DIR = Path("data/projects")
OUTPUT_DIR = Path("data/output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 如果PDF在原始位置
SOURCE_DIR = Path(r"E:\备份资料\IEOR 4524 Spring 2026-20251227T052940Z-3-001\IEOR 4524 Spring 2026")


def get_pdf_files() -> List[Path]:
    """获取所有PDF文件"""
    # 先检查项目目录
    if PROJECTS_DIR.exists():
        pdfs = list(PROJECTS_DIR.glob("*.pdf"))
        if pdfs:
            return sorted(pdfs)
    
    # 如果项目目录没有，检查源目录
    if SOURCE_DIR.exists():
        pdfs = list(SOURCE_DIR.glob("*.pdf"))
        if pdfs:
            return sorted(pdfs)
    
    return []


def extract_pdf_text(pdf_path: Path) -> str:
    """提取PDF文本"""
    try:
        import PyPDF2
        with open(pdf_path, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    except ImportError:
        try:
            import pdfplumber
            text = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text
        except ImportError:
            print("错误: 需要安装 PyPDF2 或 pdfplumber")
            return ""
    except Exception as e:
        print(f"✗ PDF解析错误: {pdf_path.name} - {str(e)}")
        return ""


def save_texts_for_ai_analysis(pdf_files: List[Path]):
    """提取所有PDF文本，保存供AI分析"""
    texts_dir = Path("data/project_texts")
    texts_dir.mkdir(parents=True, exist_ok=True)
    
    all_texts = []
    
    print(f"开始提取 {len(pdf_files)} 个PDF的文本...\n")
    
    for i, pdf_file in enumerate(pdf_files, 1):
        print(f"[{i}/{len(pdf_files)}] 处理: {pdf_file.name}")
        
        text = extract_pdf_text(pdf_file)
        if not text:
            print(f"  ✗ 无法提取文本")
            continue
        
        # 保存单个文本文件
        text_file = texts_dir / f"{pdf_file.stem}.txt"
        with open(text_file, 'w', encoding='utf-8') as f:
            f.write(text)
        
        all_texts.append({
            "序号": i,
            "PDF文件": pdf_file.name,
            "文本文件": text_file.name,
            "文本内容": text,
            "文本长度": len(text)
        })
        
        print(f"  ✓ 文本已保存 ({len(text)} 字符)")
    
    # 保存合并的文本文件
    all_texts_file = texts_dir / "all_projects_text.txt"
    with open(all_texts_file, 'w', encoding='utf-8') as f:
        for item in all_texts:
            f.write(f"\n{'='*80}\n")
            f.write(f"项目 {item['序号']}: {item['PDF文件']}\n")
            f.write(f"{'='*80}\n\n")
            f.write(item['文本内容'])
            f.write("\n\n")
    
    # 保存索引JSON
    index_file = texts_dir / "projects_index.json"
    index_data = [
        {
            "序号": item["序号"],
            "PDF文件": item["PDF文件"],
            "文本文件": item["文本文件"],
            "文本长度": item["文本长度"],
            "文本预览": item["文本内容"][:300] + "..." if len(item["文本内容"]) > 300 else item["文本内容"]
        }
        for item in all_texts
    ]
    
    with open(index_file, 'w', encoding='utf-8') as f:
        json.dump(index_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ 完成！")
    print(f"✓ 共提取 {len(all_texts)} 个项目的文本")
    print(f"✓ 合并文本文件: {all_texts_file}")
    print(f"✓ 索引文件: {index_file}")
    
    return all_texts


def export_to_excel(projects: List[Dict], output_path: Path):
    """导出到Excel"""
    df = pd.DataFrame(projects)
    
    column_order = [
        "项目编号", "项目名称", "公司名称", "所处行业", 
        "应用场景", "公司用心程度", "预期成果", 
        "技能要求", "项目描述摘要", "源文件"
    ]
    
    existing_columns = [col for col in column_order if col in df.columns]
    other_columns = [col for col in df.columns if col not in column_order]
    df = df[existing_columns + other_columns]
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='项目分析')
        
        worksheet = writer.sheets['项目分析']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        worksheet.row_dimensions[1].height = 25
        for row in range(2, len(df) + 2):
            worksheet.row_dimensions[row].height = 20
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    print(f"✓ Excel已导出: {output_path}")


def main():
    """主函数"""
    print("=" * 60)
    print("PDF项目文本提取")
    print("=" * 60)
    
    # 获取PDF文件
    pdf_files = get_pdf_files()
    
    if not pdf_files:
        print("\n未找到PDF文件！")
        print(f"请确保PDF文件在以下位置之一：")
        print(f"1. {PROJECTS_DIR}")
        print(f"2. {SOURCE_DIR}")
        print("\n或者运行 copy_pdfs.bat 来复制文件")
        return
    
    print(f"\n找到 {len(pdf_files)} 个PDF文件")
    
    # 提取文本
    all_texts = save_texts_for_ai_analysis(pdf_files)
    
    if all_texts:
        print("\n" + "="*60)
        print("文本提取完成！")
        print("="*60)
        print("\n下一步：")
        print("现在可以让我（AI助手）开始分析这些项目了！")
        print("我会读取 data/project_texts/all_projects_text.txt 文件")
        print("然后逐个分析每个项目并提取关键信息。")


if __name__ == "__main__":
    main()


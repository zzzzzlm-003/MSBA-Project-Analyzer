"""
项目分析系统 - 从Google Drive下载PDF，AI提取关键信息，导出Excel
"""

import os
import json
import re
import requests
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import openai
from dotenv import load_dotenv

load_dotenv()

# 配置
PROJECTS_DIR = Path("data/projects")
PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = Path("data/output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# OpenAI配置
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    print("警告: 未找到OPENAI_API_KEY环境变量，请设置后使用AI提取功能")

class GoogleDriveDownloader:
    """从Google Drive共享链接下载PDF文件"""
    
    @staticmethod
    def extract_file_id(share_link: str) -> Optional[str]:
        """从Google Drive共享链接中提取文件ID"""
        patterns = [
            r'/file/d/([a-zA-Z0-9_-]+)',
            r'id=([a-zA-Z0-9_-]+)',
            r'/([a-zA-Z0-9_-]{25,})',  # Google Drive文件ID通常是25+字符
        ]
        
        for pattern in patterns:
            match = re.search(pattern, share_link)
            if match:
                return match.group(1)
        return None
    
    @staticmethod
    def download_file(file_id: str, output_path: Path, filename: str = None) -> bool:
        """下载Google Drive文件"""
        # 使用export格式下载（适用于PDF）
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        
        try:
            # 首先获取确认页面（大文件需要确认）
            session = requests.Session()
            response = session.get(f"https://drive.google.com/uc?id={file_id}", stream=True)
            
            # 检查是否需要确认
            if 'download_warning' in response.url or 'confirm' in response.url:
                # 提取确认token
                confirm_match = re.search(r'confirm=([^&]+)', response.url)
                if confirm_match:
                    confirm_token = confirm_match.group(1)
                    download_url = f"https://drive.google.com/uc?export=download&id={file_id}&confirm={confirm_token}"
                    response = session.get(download_url, stream=True)
            
            # 如果响应是HTML（错误页面），尝试直接下载
            if response.headers.get('Content-Type', '').startswith('text/html'):
                # 尝试使用不同的下载方式
                download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
                response = requests.get(download_url, stream=True, allow_redirects=True)
            
            if response.status_code == 200:
                output_path.parent.mkdir(parents=True, exist_ok=True)
                with open(output_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                print(f"✓ 下载成功: {output_path.name}")
                return True
            else:
                print(f"✗ 下载失败: {response.status_code} - {output_path.name}")
                return False
                
        except Exception as e:
            print(f"✗ 下载错误: {output_path.name} - {str(e)}")
            return False
    
    @staticmethod
    def download_from_folder_link(folder_link: str, output_dir: Path) -> List[str]:
        """从Google Drive文件夹链接下载所有PDF文件"""
        # 注意：这需要Google Drive API，或者手动提供文件列表
        # 这里提供一个简化的方法：从文件夹页面提取文件ID
        print("提示: 文件夹批量下载需要Google Drive API或手动提供文件链接列表")
        return []


class PDFExtractor:
    """提取PDF文本内容"""
    
    @staticmethod
    def extract_text(pdf_path: Path) -> str:
        """从PDF文件中提取文本"""
        try:
            # 使用PyPDF2或pdfplumber作为替代
            # 这里先尝试使用PyPDF2
            try:
                import PyPDF2
                with open(pdf_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
                return text
            except ImportError:
                # 如果PyPDF2不可用，尝试pdfplumber
                try:
                    import pdfplumber
                    text = ""
                    with pdfplumber.open(pdf_path) as pdf:
                        for page in pdf.pages:
                            text += page.extract_text() + "\n"
                    return text
                except ImportError:
                    print("错误: 需要安装 PyPDF2 或 pdfplumber")
                    return ""
        except Exception as e:
            print(f"✗ PDF解析错误: {pdf_path.name} - {str(e)}")
            return ""


class AIInfoExtractor:
    """使用AI提取项目关键信息"""
    
    def __init__(self, api_key: str = None):
        self.api_key = api_key or OPENAI_API_KEY
        if not self.api_key:
            raise ValueError("需要设置OPENAI_API_KEY")
        openai.api_key = self.api_key
    
    def extract_project_info(self, pdf_text: str, filename: str) -> Dict:
        """使用AI提取项目信息"""
        
        # 限制文本长度
        max_length = 8000
        if len(pdf_text) > max_length:
            pdf_text = pdf_text[:max_length] + "...[文本已截断]"
        
        prompt = f"""请从以下项目文档中提取关键信息。文档内容：

{pdf_text}

请提取以下信息，并以JSON格式返回：
{{
  "项目名称": "项目标题或名称",
  "所处行业": "如：金融、医疗、教育、科技、零售等",
  "应用场景": "项目的具体应用场景和用途，详细描述",
  "公司用心程度": "根据文档的详细程度、完整性、专业性评分，1-10分，并说明理由",
  "预期成果": "项目预期交付的成果或产出",
  "项目编号": "如果有项目编号或ID",
  "公司名称": "合作公司或组织名称",
  "技能要求": "所需技能和技术栈",
  "项目描述摘要": "项目简要描述（100字以内）"
}}

只返回JSON，不要其他文字。"""
        
        try:
            # 使用OpenAI API
            response = openai.ChatCompletion.create(
                model="gpt-4o-mini",  # 或使用 "gpt-4" 获得更好效果
                messages=[
                    {"role": "system", "content": "你是一个专业的项目分析助手，擅长从项目文档中提取结构化信息。请只返回JSON格式，不要添加任何解释文字。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=1000
            )
            
            result_text = response.choices[0].message.content.strip()
            
            # 尝试解析JSON
            # 移除可能的markdown代码块标记
            if result_text.startswith("```json"):
                result_text = result_text[7:]
            if result_text.startswith("```"):
                result_text = result_text[3:]
            if result_text.endswith("```"):
                result_text = result_text[:-3]
            result_text = result_text.strip()
            
            info = json.loads(result_text)
            info["源文件"] = filename
            return info
            
        except json.JSONDecodeError as e:
            print(f"✗ JSON解析错误: {filename}")
            print(f"AI返回内容: {result_text[:200]}")
            return {
                "项目名称": "解析失败",
                "所处行业": "未知",
                "应用场景": "无法提取",
                "公司用心程度": "0 - JSON解析失败",
                "预期成果": "无法提取",
                "源文件": filename
            }
        except Exception as e:
            print(f"✗ AI提取错误: {filename} - {str(e)}")
            return {
                "项目名称": "提取失败",
                "所处行业": "未知",
                "应用场景": "无法提取",
                "公司用心程度": "0 - 提取失败",
                "预期成果": "无法提取",
                "源文件": filename
            }


class ExcelExporter:
    """导出数据到Excel"""
    
    @staticmethod
    def export_to_excel(projects: List[Dict], output_path: Path):
        """导出项目数据到Excel文件"""
        df = pd.DataFrame(projects)
        
        # 重新排列列的顺序，重要信息在前
        column_order = [
            "项目编号", "项目名称", "公司名称", "所处行业", 
            "应用场景", "公司用心程度", "预期成果", 
            "技能要求", "项目描述摘要", "源文件"
        ]
        
        # 只保留存在的列
        existing_columns = [col for col in column_order if col in df.columns]
        other_columns = [col for col in df.columns if col not in column_order]
        df = df[existing_columns + other_columns]
        
        # 使用openpyxl创建格式化的Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='项目分析')
            
            # 获取工作表
            worksheet = writer.sheets['项目分析']
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 设置列宽，但不超过50
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 设置行高
            worksheet.row_dimensions[1].height = 25
            for row in range(2, len(df) + 2):
                worksheet.row_dimensions[row].height = 20
            
            # 设置文本换行
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        print(f"✓ Excel文件已导出: {output_path}")
        return df


class ProjectAnalyzer:
    """项目分析主程序"""
    
    def __init__(self):
        self.downloader = GoogleDriveDownloader()
        self.extractor = PDFExtractor()
        self.ai_extractor = None
        if OPENAI_API_KEY:
            try:
                self.ai_extractor = AIInfoExtractor()
            except Exception as e:
                print(f"警告: AI提取器初始化失败 - {str(e)}")
        self.exporter = ExcelExporter()
    
    def download_pdfs_from_links(self, drive_links: List[str]) -> List[Path]:
        """从Google Drive链接列表下载PDF"""
        downloaded_files = []
        
        for i, link in enumerate(drive_links, 1):
            print(f"\n[{i}/{len(drive_links)}] 处理链接: {link}")
            
            file_id = self.downloader.extract_file_id(link)
            if not file_id:
                print(f"✗ 无法提取文件ID: {link}")
                continue
            
            # 生成文件名
            filename = f"project_{i:03d}.pdf"
            output_path = PROJECTS_DIR / filename
            
            if self.downloader.download_file(file_id, output_path):
                downloaded_files.append(output_path)
        
        return downloaded_files
    
    def analyze_projects(self, pdf_files: List[Path] = None) -> List[Dict]:
        """分析所有PDF项目"""
        if pdf_files is None:
            # 从目录中读取所有PDF
            pdf_files = list(PROJECTS_DIR.glob("*.pdf"))
        
        if not pdf_files:
            print("未找到PDF文件")
            return []
        
        projects = []
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"\n[{i}/{len(pdf_files)}] 分析: {pdf_file.name}")
            
            # 提取PDF文本
            text = self.extractor.extract_text(pdf_file)
            if not text:
                print(f"✗ 无法提取文本: {pdf_file.name}")
                continue
            
            # AI提取信息
            if self.ai_extractor:
                info = self.ai_extractor.extract_project_info(text, pdf_file.name)
                projects.append(info)
                print(f"✓ 提取完成: {info.get('项目名称', '未知')}")
            else:
                print("✗ AI提取器未初始化，跳过")
        
        return projects
    
    def export_results(self, projects: List[Dict], format: str = "excel"):
        """导出结果"""
        if format == "excel":
            timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
            output_path = OUTPUT_DIR / f"项目分析_{timestamp}.xlsx"
            df = self.exporter.export_to_excel(projects, output_path)
            return df
        elif format == "json":
            output_path = OUTPUT_DIR / f"项目分析_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(projects, f, ensure_ascii=False, indent=2)
            print(f"✓ JSON文件已导出: {output_path}")
            return projects


def load_config() -> Dict:
    """加载配置文件"""
    config_path = Path("project_analyzer_config.json")
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"警告: 配置文件读取失败 - {str(e)}")
    return {}


def main():
    """主函数"""
    analyzer = ProjectAnalyzer()
    config = load_config()
    
    print("=" * 60)
    print("项目分析系统")
    print("=" * 60)
    
    # 方式1: 从Google Drive链接列表下载
    print("\n请选择操作模式:")
    print("1. 从Google Drive链接列表下载并分析")
    print("2. 从配置文件读取链接并分析")
    print("3. 分析本地已有的PDF文件")
    
    choice = input("\n请输入选择 (1/2/3): ").strip()
    
    links = []
    
    if choice == "1":
        print("\n请输入Google Drive链接（每行一个，输入空行结束）:")
        while True:
            link = input().strip()
            if not link:
                break
            links.append(link)
    elif choice == "2":
        # 从配置文件读取
        if "google_drive_links" in config:
            links = [link for link in config["google_drive_links"] 
                    if link and not link.startswith("在此处") and not link.startswith("例如")]
            print(f"\n从配置文件读取到 {len(links)} 个链接")
        else:
            print("配置文件中没有找到google_drive_links")
            return
    
    if choice in ["1", "2"] and links:
        pdf_files = analyzer.download_pdfs_from_links(links)
        if pdf_files:
            projects = analyzer.analyze_projects(pdf_files)
        else:
            print("未成功下载任何PDF文件")
            return
    elif choice == "3":
        projects = analyzer.analyze_projects()
    else:
        print("未提供链接或选择无效")
        return
    
    if projects:
        # 导出Excel
        df = analyzer.export_results(projects, "excel")
        # 同时导出JSON（用于Notion导入）
        analyzer.export_results(projects, "json")
        print(f"\n✓ 共分析 {len(projects)} 个项目")
        print(f"✓ 结果已保存到: {OUTPUT_DIR}")
        print(f"\n提示: 如需导入Notion，运行 python notion_integration.py")
    else:
        print("\n✗ 未找到可分析的项目")


if __name__ == "__main__":
    main()
